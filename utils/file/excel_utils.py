#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File  : excel_utils.py
# @Author: 尘昊2259
# @Date  : 2022/2/12 14:13
# @Desc  : excel的操作
import traceback

from openpyxl import load_workbook
from utils.file.path_utils import PathBase
from utils.log.loguru_utils import Logger


class ExcelUtils:
    def __init__(self):
        path_object = PathBase()
        self.date_path = path_object.join_path(path_object
                                               .get_path(path_object
                                                         .get_path(path_object.path)), "data")
        self.path = ".xlsx"

    def read_excel(self, filename, sheetname):
        """

        :param filepath: 读取的文件路径
        :param sheetname: 需要读取的excel名字
        :return:
        """
        try:
            # 打开excel
            openexcel = load_workbook(f"{self.date_path}/{filename}{self.path}")
            # 指定excel要读取的sheet
            sheet = openexcel[sheetname]
            # 获取最大行
            maxrow = sheet.max_row
            # 获取最大列
            maxlist = sheet.max_column

            alldata = []
            # 遍历行
            for r in range(2, maxrow + 1):
                # 存取读取的结果
                read_result = []
                # 遍历列
                for l in range(1, maxlist + 1):
                    read_result.append(sheet.cell(r, l).value)

                # 用于存储处理后的数据
                read_result2 = []
                # 遍历结果，去掉换行符和None，不要这些数据
                for result in read_result:
                    if result is not None:  # 过滤None
                        result1 = str(result)  # 将结果转换为字符串
                        result2 = result1.replace("\n", "")  # 将换行符替换为空
                        read_result2.append(result2)

                alldata.append(read_result2)

            return alldata
        except Exception as e:
            Logger().error(traceback.format_exc())

    def extract_test_data(self, filename, sheetname):
        # 存取测试数据和预期结果
        alldatas = []

        try:
            # 获取excel读取的数据
            result = ExcelUtils().read_excel(filename, sheetname)

            for test_data in result:
                # 预期结果
                expected = test_data[4]
                # 取出测试数据，并转换为字典
                test_data2 = eval(test_data[3])

                # 存储遍历字典后的value和预期结果
                value_expected = []
                # 遍历字典格式的测试数据，取出value
                if isinstance(test_data2, dict):
                    for values in test_data2.values():
                        # 追加value
                        value_expected.append(values)
                    # 追加预期结果
                    value_expected.append(expected)

                # 追加最终的数据
                alldatas.append(value_expected)
            return alldatas
        except Exception as e:
            Logger().error(traceback.format_exc())


if __name__ == '__main__':
    a = ExcelUtils().extract_test_data("addstudent", "新增学生")
    print(a)
